from ScrumPy.Util import Set
from importlib import reload
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
import BuildLP_ref

'''
new function to set each biomass producing reaction to -1.0
'''
def MinimalBM(m):
    bm_list = list(filter(lambda s:'_bm_tx' in s, m.sm.cnames))
    bm_reac_dict ={} 
    for bm in bm_list:
        bm_reac_dict[bm] = (1.0,None)
    return bm_reac_dict

'''
new function to compare LP solutions
'''
def CompareSol(sol1,sol2):
    only_1 = '\n'+'#'*100 + '\n'+'following reactions were present only in the first model \n'+'#'*100 + '\n\n'
    only_2 = '\n'+'#'*100 + '\n'+'following reactions were present only in the second model \n'+'#'*100 + '\n\n'
    difference = '\n'+'#'*100 + '\n'+'following reactions were significantly different in both models \n'+'#'*100 + '\n\n'
    same = '\n'+'#'*100 + '\n'+'following reactions were similar in both models \n'+'#'*100 + '\n\n'
    new_sol = {}
    for r in Set.Complement(sol1,sol2):
        only_1 += f'{r}\t{sol1[r]} \n'
    for r in Set.Intersect(sol1,sol2):
        ratio = abs(sol1[r]/sol2[r])
        diff = abs(sol1[r]-sol2[r])
        if ratio > 1.001 or ratio < 0.999 :
            if sol1[r] > sol2[r]:
                difference += f'{r}\t{(sol1[r],sol2[r])}\t{-round(diff*100/abs(sol1[r]),3)}%\n'
            else:
                difference += f'{r}\t{(sol1[r],sol2[r])}\t{+round(diff*100/abs(sol1[r]),3)}%\n'
        else:
            same += f'{r}\t{(sol1[r],sol2[r])}\n'
    for r in Set.Complement(sol2,sol1):
        only_2 += f'{r}\t{sol2[r]} \n'
    while True:
        main = input('\nSelect an option or press another button to exit\n\n1. Reactions present only in the first model\n2. Reactions present only in the second model\n3. Significantly different reactions\n4. Similar reactions\n>')
        if main=='1':
            print(only_1)
        elif main=='2':
            print(only_2)
        elif main=='3':
            print(difference)
        elif main=='4':
            print(same)
        else:
            break

'''
new function to perform reaction scan
'''

def Scan(lp, r = None, upper=None, lower=None): 
    if r==None:
        r= input('input the reaction to scan\n>')
    if upper==None:
        upper = float(input(f'rate upperbound for {r}\n>'))
    if lower==None:
        lower = float(input(f'rate lowerbound for {r}\n>')) 
    step = float(input('steps\n>'))
    step_len = (upper-lower)/step
    constraints = {}
    h = 0
    df = {}
    j = lower
    while j <= upper:
        print(j)
        lp.SetFixedFlux({r:j})
        lp.Solve()
        sol = lp.GetPrimSol()
        for k in sol:
            if k in df:
                df[k].append(sol[k])
            else:
                df[k] = [0 for i in range(h)]
                df[k].append(sol[k])
        h += 1
        for k in df:
            if len(df[k])!=h:
                df[k].append(0) #in case some reaction values drop to 0
        j += step_len
    
    #remove undefined rows --> rows with all 0 values
    empty = []
    for i in range(h): #each row should have the same length
        check = True
        for k in df:
            if df[k][i] != 0:
                check = False
        if check:
            empty.append(i)
            
    for i in empty:
        for k in df:
            df[k].pop(i)
    
    
    #write the dataframe into an excel file
    df1=pd.DataFrame(df)
    df1.to_excel('scan.xlsx')
    print(f'total reactions are stored into scan.xlsx\nundefined solutions {empty} are removed')
    #remove reactions with the same value in row --> reactions not changing flux value
    newd = {}
    for k in df:
        res = []
        for v in df[k]:
            val = round(v,9) #otherwise some equal values are considered the different for extremely small differences
            if val not in res:
                res.append(val)
        if len(res) != 1:
            newd[k] = df[k]
    df2=pd.DataFrame(newd)
    df2.to_excel('unstable.xlsx')
    print('variable reactions are stored into unstable.xlsx')
    
    #return df
    #following code to be implemented --> write the dataframe into a file

'''
useless function because repeating the same FBA problem in the same ScrumPy window gives you the same results


def Repeat(m,n): #for now it works only for the reference (BuildLP_ref)
    tot_b = []
    h = 0
    df = {}
    for i in range(n):
        lp = BuildLP_ref.BoundedUptakeLP(m)
        lp.Solve()
        sol = lp.GetPrimSol()
        for k in sol:
            if k in df:
                df[k].append(sol[k])
            else:
                df[k] = [0 for i in range(h)]
                df[k].append(sol[k])
        h += 1
        for k in df:
            if len(df[k])!=h:
                df[k].append(0) #in case some reaction values drop to 0
    
    df1=pd.DataFrame(df)
    df1.to_excel('repeatition.xlsx')
    print('results are stored into repeatition.xlsx')
    return df

'''



def excel():
    '''
    workbook=Workbook()
    sheet=workbook.active
    sheet["A1"] = "hello"
    sheet["B1"] = "world!"
    workbook.save(filename="hello_world.xlsx")
    '''
    '''
    workbook = load_workbook(filename='hello_world.xlsx')
    sheet = workbook.active
    
    sheet["C1"] = 'writing ;)'
    '''
    
        
 